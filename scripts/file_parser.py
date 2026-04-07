# -*- coding: utf-8 -*-
"""
非 Markdown 文件解析器
将 txt / csv / xlsx / docx 转换为飞书 Block 格式
接口与 MarkdownParser 兼容：blocks, pending_images, pending_tables
"""

import csv
import io
from pathlib import Path
from typing import Dict, List, Any, Tuple


class FileParser:
    """将 txt/csv/xlsx/docx 解析为飞书 Block 列表"""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.pending_images: List[Tuple[int, str, bool]] = []   # 与 MarkdownParser 接口一致，始终为空
        self.pending_tables: List[Tuple[int, List[List[str]]]] = []

    def parse(self) -> List[Dict[str, Any]]:
        """根据扩展名分发到对应解析方法"""
        ext = self.file_path.suffix.lower()
        if ext == ".txt":
            return self._parse_txt()
        elif ext == ".csv":
            return self._parse_csv()
        elif ext == ".xlsx":
            return self._parse_xlsx()
        elif ext == ".docx":
            return self._parse_docx()
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

    # ------------------------------------------------------------------ #
    # TXT
    # ------------------------------------------------------------------ #
    def _parse_txt(self) -> List[Dict[str, Any]]:
        with open(self.file_path, "r", encoding="utf-8") as f:
            content = f.read()

        blocks = []
        paragraphs = content.split("\n\n")   # 空行分段

        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            # 简单识别"## 标题"格式（兼容部分带格式的 txt）
            if para.startswith("## "):
                blocks.append(self._heading(para[3:].strip(), 2))
            elif para.startswith("# "):
                blocks.append(self._heading(para[2:].strip(), 1))
            else:
                # 多行文本合并为一个段落
                text = para.replace("\n", " ")
                blocks.append(self._text(text))

        return blocks

    # ------------------------------------------------------------------ #
    # CSV
    # ------------------------------------------------------------------ #
    def _parse_csv(self) -> List[Dict[str, Any]]:
        with open(self.file_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            rows = [row for row in reader if any(cell.strip() for cell in row)]

        if not rows:
            return [self._text("（空文件）")]

        blocks = []
        # 用文件名作为标题
        blocks.append(self._heading(self.file_path.stem, 1))
        # 注册表格到 pending_tables
        placeholder_idx = len(blocks)
        self.pending_tables.append((placeholder_idx, rows))
        blocks.append(self._table_placeholder())

        return blocks

    # ------------------------------------------------------------------ #
    # XLSX
    # ------------------------------------------------------------------ #
    def _parse_xlsx(self) -> List[Dict[str, Any]]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("请先安装 openpyxl: pip install openpyxl")

        wb = openpyxl.load_workbook(self.file_path, data_only=True)
        blocks = []
        # 文件名作为总标题
        blocks.append(self._heading(self.file_path.stem, 1))

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                str_row = [str(cell) if cell is not None else "" for cell in row]
                if any(c.strip() for c in str_row):
                    rows.append(str_row)

            if not rows:
                continue

            # Sheet 名称作为二级标题
            blocks.append(self._heading(sheet_name, 2))
            placeholder_idx = len(blocks)
            self.pending_tables.append((placeholder_idx, rows))
            blocks.append(self._table_placeholder())

        return blocks

    # ------------------------------------------------------------------ #
    # DOCX
    # ------------------------------------------------------------------ #
    def _parse_docx(self) -> List[Dict[str, Any]]:
        try:
            from docx import Document
            from docx.oxml.ns import qn
        except ImportError:
            raise ImportError("请先安装 python-docx: pip install python-docx")

        doc = Document(self.file_path)
        blocks = []

        # 遍历文档正文元素（段落 + 表格，保持顺序）
        from docx.table import Table as DocxTable
        from docx.text.paragraph import Paragraph as DocxParagraph

        for element in doc.element.body:
            tag = element.tag.split("}")[-1] if "}" in element.tag else element.tag

            if tag == "p":
                para = DocxParagraph(element, doc)
                text = para.text.strip()
                if not text:
                    continue
                style = para.style.name if para.style else ""
                if style.startswith("Heading 1") or style.startswith("标题 1"):
                    blocks.append(self._heading(text, 1))
                elif style.startswith("Heading 2") or style.startswith("标题 2"):
                    blocks.append(self._heading(text, 2))
                elif style.startswith("Heading 3") or style.startswith("标题 3"):
                    blocks.append(self._heading(text, 3))
                else:
                    blocks.append(self._text(text))

            elif tag == "tbl":
                table = DocxTable(element, doc)
                rows = []
                for row in table.rows:
                    rows.append([cell.text.strip() for cell in row.cells])
                if rows:
                    placeholder_idx = len(blocks)
                    self.pending_tables.append((placeholder_idx, rows))
                    blocks.append(self._table_placeholder())

        return blocks

    # ------------------------------------------------------------------ #
    # Block 构造辅助
    # ------------------------------------------------------------------ #
    def _heading(self, text: str, level: int) -> Dict[str, Any]:
        block_type = 2 + level   # level1=3, level2=4, level3=5
        return {
            "block_type": block_type,
            f"heading{level}": {
                "elements": [{"text_run": {"content": text}}],
                "style": {}
            }
        }

    def _text(self, text: str) -> Dict[str, Any]:
        return {
            "block_type": 2,
            "text": {
                "elements": [{"text_run": {"content": text}}],
                "style": {}
            }
        }

    def _table_placeholder(self) -> Dict[str, Any]:
        """表格占位符，实际写入由 writer.py 中的 _write_content_with_images 处理"""
        return {"block_type": 2, "text": {"elements": [{"text_run": {"content": ""}}], "style": {}}}
